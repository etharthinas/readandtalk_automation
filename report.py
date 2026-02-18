from crawling import StudentInfo
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.label import DataLabelList
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference

grey_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
blue_fill = PatternFill(fill_type="solid", start_color="4472C4", end_color="4472C4")
orange_fill = PatternFill(fill_type="solid", start_color="ED7D31", end_color="ED7D31")

center = Alignment(horizontal="center", vertical="center", wrap_text=True)

COMMA_FORMAT_INTEGER = "#,##0"

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

thick_side = Side(style="thick")


def create_learning_report(student_info: StudentInfo, comments: str, dir="/"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Report"

    for i in ["A", "B", "C", "D", "E", "F", "G"]:
        ws.column_dimensions[i].width = 13

    # ---------- HEADER ----------
    ws.merge_cells("A1:G1")
    ws["A1"] = "Individual Learning Reports"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = center

    # --------- Basic Info ---------
    ws["A2"] = "학습 기간"
    apply_grey(ws["A2"])

    ws.merge_cells("B2:D2")
    set_default(
        ws["B2"],
        student_info.basic_info.time_start[:4]
        + "년 "
        + student_info.basic_info.time_start[4:]
        + "월 ~ "
        + student_info.basic_info.time_end[:4]
        + "년 "
        + student_info.basic_info.time_end[4:]
        + "월",
    )

    ws["E2"] = "예상 Lexile"
    apply_grey(ws["E2"])

    ws.merge_cells("F2:G2")
    set_default(ws["F2"], str(student_info.basic_info.lexile) + "L")

    ws["A3"] = "이름"
    apply_grey(ws["A3"])

    set_default(ws["B3"], student_info.basic_info.name)

    ws["C3"] = "학교"
    apply_grey(ws["C3"])

    set_default(ws["D3"], student_info.basic_info.school)

    ws["E3"] = "학년"
    apply_grey(ws["E3"])

    ws.merge_cells("F3:G3")
    set_default(ws["F3"], student_info.basic_info.grade)

    ws["A4"] = "수업차수"
    apply_grey(ws["A4"])

    set_default(ws["B4"], student_info.basic_info.count)

    ws["C4"] = "학습단계"
    apply_grey(ws["C4"])

    ws.merge_cells("D4:G4")
    set_default(ws["D4"], student_info.basic_info.level)

    apply_thick_border(ws, row_start=2, row_end=4, column_start=1, column_end=7)

    # ----- Sections ------
    create_section(
        ws,
        title="원서 학습량",
        headings=[
            "정독(권)",
            "다독(권)",
            "인문고전(권)",
            "당분기권)",
            "전분기(권)",
            "총 학습량(권)",
        ],
        data=[
            student_info.book_info[0].intensive,
            student_info.book_info[0].extensive,
            student_info.book_info[0].classics,
            student_info.book_info[0].curr_month,
            student_info.book_info[1].curr_month,
            student_info.book_info[0].total,
        ],
        start_row=6,
        flag=True,
    )

    create_section(
        ws,
        title="Word 학습량\n정답률",
        headings=[
            "당분기(개)",
            "정답률(%)",
            "전분기(개)",
            "정답률(%)",
            "총학습량(개)",
            "정답률(%)",
        ],
        data=[
            student_info.word_info[0].curr_count,
            student_info.word_info[0].curr_rate,
            student_info.word_info[1].curr_count,
            student_info.word_info[1].curr_rate,
            student_info.word_info[0].total_count,
            student_info.word_info[0].total_rate,
        ],
        start_row=10,
    )

    create_section(
        ws,
        title="Puzzle 학습량\n정답률",
        headings=[
            "당분기(문장)",
            "정답률(%)",
            "전분기(문장)",
            "정답률(%)",
            "총학습량(문장)",
            "정답률(%)",
        ],
        data=[
            student_info.puzzle_info[0].curr_count,
            student_info.puzzle_info[0].curr_rate,
            student_info.puzzle_info[1].curr_count,
            student_info.puzzle_info[1].curr_rate,
            student_info.puzzle_info[0].total_count,
            student_info.puzzle_info[0].total_rate,
        ],
        start_row=14,
    )

    create_section(
        ws,
        title="Dictation 학습량\n정답률",
        headings=[
            "당분기(문장)",
            "정답률(%)",
            "전분기(문장)",
            "정답률(%)",
            "총학습량(문장)",
            "정답률(%)",
        ],
        data=[
            student_info.dictation_info[0].curr_count,
            student_info.dictation_info[0].curr_rate,
            student_info.dictation_info[1].curr_count,
            student_info.dictation_info[1].curr_rate,
            student_info.dictation_info[0].total_count,
            student_info.dictation_info[0].total_rate,
        ],
        start_row=18,
    )

    create_section(
        ws,
        title="Writing 학습량\n정답률",
        headings=[
            "당분기(문장)",
            "정답률(%)",
            "전분기(문장)",
            "정답률(%)",
            "총학습량(문장)",
            "정답률(%)",
        ],
        data=[
            student_info.writing_info[0].curr_count,
            student_info.writing_info[0].curr_rate,
            student_info.writing_info[1].curr_count,
            student_info.writing_info[1].curr_rate,
            student_info.writing_info[0].total_count,
            student_info.writing_info[0].total_rate,
        ],
        start_row=22,
    )

    create_section(
        ws,
        title="Quiz 학습량\n정답률",
        headings=[
            "당분기(문제)",
            "정답률(%)",
            "전분기(문제)",
            "정답률(%)",
            "총학습량(문제)",
            "정답률(%)",
        ],
        data=[
            student_info.quiz_info[0].curr_count,
            student_info.quiz_info[0].curr_rate,
            student_info.quiz_info[1].curr_count,
            student_info.quiz_info[1].curr_rate,
            student_info.quiz_info[0].total_count,
            student_info.quiz_info[0].total_rate,
        ],
        start_row=26,
    )

    # ----- Footer -------
    ws.merge_cells("A29:G29")
    apply_grey(ws["A29"])
    ws["A29"].font = Font(bold=True, size=18)
    ws["A29"].alignment = center
    ws["A29"] = "Teacher's Comments"
    ws.row_dimensions[29].height = 70
    apply_thick_border(ws, row_start=29, row_end=29, column_start=1, column_end=7)

    ws.merge_cells("A30:G35")
    set_default(ws["A30"], comments)
    ws["A30"].alignment = Alignment(
        horizontal="left", vertical="center", wrap_text=True
    )

    wb.save(
        dir
        / f"{student_info.basic_info.time_start[2:]}_{student_info.basic_info.time_end[2:]}_{student_info.basic_info.name}_통신문.xlsx"
    )


def create_section(
    ws, title: str, headings: list[str], data: list[int], start_row: int, flag=False
):
    if len(headings) != 6 or len(data) != 6:
        raise ValueError
    ws.merge_cells("A{s}:A{e}".format(s=start_row, e=start_row + 1))
    ws[f"A{start_row}"] = title
    apply_grey(ws[f"A{start_row}"])

    for i, col in enumerate(["B", "C", "D", "E", "F", "G"]):
        ws[f"{col}{start_row}"] = headings[i]
        apply_grey(ws[f"{col}{start_row}"])
        ws[f"{col}{start_row+1}"] = data[i]
        if (col == "C" and not flag) or (col == "E" and flag):
            apply_blue(ws[f"{col}{start_row+1}"])
        elif (col == "E" and not flag) or (col == "F" and flag):
            apply_orange(ws[f"{col}{start_row+1}"])
        else:
            set_default(ws[f"{col}{start_row+1}"], data[i])
            if col == "F":
                ws[f"{col}{start_row+1}"].number_format = COMMA_FORMAT_INTEGER
            if col == "G" and flag:
                ws[f"{col}{start_row+1}"].number_format = COMMA_FORMAT_INTEGER

    apply_thick_border(
        ws, row_start=start_row, row_end=start_row + 1, column_start=1, column_end=7
    )

    ws.row_dimensions[start_row + 2].height = 70
    chart = BarChart()
    chart.type = "bar"
    chart.height = 2
    chart.style = 2
    chart.width = 18
    chart.legend = None

    data_curr = Reference(
        ws,
        min_col=3 if not flag else 5,
        min_row=start_row + 1,
        max_col=3 if not flag else 5,
        max_row=start_row + 1,
    )
    data_prev = Reference(
        ws,
        min_col=5 if not flag else 6,
        min_row=start_row + 1,
        max_col=5 if not flag else 6,
        max_row=start_row + 1,
    )

    chart.add_data(data_prev, titles_from_data=False)
    chart.add_data(data_curr, titles_from_data=False)

    chart.series[0].graphicalProperties.solidFill = "ED7D31"
    chart.series[1].graphicalProperties.solidFill = "4472C4"

    chart.y_axis.min = 0
    chart.y_axis.max = 100 if not flag else 25
    chart.y_axis.majorGridlines = ChartLines()

    # chart.x_axis.crosses = "min"
    chart.y_axis.majorUnit = 20 if not flag else 5
    chart.y_axis.axPos = "b"

    chart.dataLabels = DataLabelList()
    chart.dataLabels.showVal = True  # show the numeric value
    chart.dataLabels.showLegendKey = False  # do NOT show series name
    chart.dataLabels.showCatName = False  # do NOT show category name
    chart.dataLabels.showSerName = False  # do NOT show series name

    ws.add_chart(chart, f"A{start_row+2}")


def set_default(cell, value):
    cell.font = Font(size=10)
    cell.alignment = center
    cell.border = thin_border
    cell.value = value


def apply_grey(cell):
    cell.fill = grey_fill
    cell.font = Font(size=10)
    cell.alignment = center
    cell.border = thin_border


def apply_blue(cell):
    cell.fill = blue_fill
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.alignment = center
    cell.border = thin_border


def apply_orange(cell):
    cell.fill = orange_fill
    cell.font = Font(bold=True, size=10, color="FFFFFF")
    cell.alignment = center
    cell.border = thin_border


def apply_thick_border(
    ws, row_start: int, row_end: int, column_start: int, column_end: int
):
    for r in range(row_start, row_end + 1):
        for c in range(column_start, column_end + 1):

            current_cell = ws.cell(row=r, column=c)

            # Create a new Border object for each cell
            # We MUST do this, or we will modify the border of other cells
            new_border = Border()

            # Apply TOP border
            if r == row_start:
                new_border.top = thick_side

            # Apply BOTTOM border
            if r == row_end:
                new_border.bottom = thick_side

            # Apply LEFT border
            if c == column_start:
                new_border.left = thick_side

            # Apply RIGHT border
            if c == column_end:
                new_border.right = thick_side

            # Assign the new border to the cell
            current_cell.border = new_border
